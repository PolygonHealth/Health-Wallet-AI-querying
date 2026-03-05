"""Integration tests for POST /api/v1/benchmark."""

import pytest


@pytest.mark.asyncio
async def test_returns_excel_with_correct_content_type(async_client_seeded):
    resp = await async_client_seeded.post(
        "/api/v1/benchmark",
        json={
            "queries": [
                {
                    "patient_id": "patient-1",
                    "query": "What conditions?",
                    "expected_answer": "",
                },
            ],
            "strategies": ["naive_dump"],
            "models": ["mock"],
        },
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers.get("content-type", "")


@pytest.mark.asyncio
async def test_excel_has_correct_columns(async_client_seeded):
    resp = await async_client_seeded.post(
        "/api/v1/benchmark",
        json={
            "queries": [
                {"patient_id": "patient-1", "query": "test", "expected_answer": ""},
            ],
            "strategies": ["naive_dump"],
            "models": ["mock"],
        },
    )
    assert resp.status_code == 200
    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "Query" in headers
    assert "Strategy" in headers
    assert "Model" in headers
    assert "Response" in headers


@pytest.mark.asyncio
async def test_empty_queries_returns_empty_excel(async_client_seeded):
    resp = await async_client_seeded.post(
        "/api/v1/benchmark",
        json={"queries": [], "strategies": ["naive_dump"], "models": ["mock"]},
    )
    assert resp.status_code == 200
    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row >= 1  # headers row
